from tkinter import *

ventana = Tk()
ventana.geometry("920x500")
ventana.title("Cálculo de acero en losas")
ventana.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico')

def calcular():
    fc = float (txt2.get())
    fy = float (txt3.get())
    b = float (txt4.get())
    h = float (txt5.get())
    r = float (txt6.get())
    pmin = float (txt7.get())
    fiv = float (txt8.get())  
    Mux = float (txt9.get())
    Muy = float (txt10.get())
    Vux = float (txt11.get())
    Vuy = float (txt12.get())
    Nbx = float (txt13.get())
    Nby = float (txt14.get())
    d = h - r
    print ('')
    print ('- DISEÑO LOSAS -')
    print ('Datos iniciales:')
    print (' fc   =', fc, '_ MPa')
    print (' fy   =', fy, '_ MPa')
    print (' b    =', b, '_ cm')
    print (' h    =', h, '_ cm')
    print (" r    = %.1f" %r, '_ cm')
    print (' pmin =', pmin)
    print (' øv   =', fiv)
    print (' Mux   =', Mux, '_ kN.cm')
    print (' Muy   =', Muy, '_ kN.cm')
    print (' Vux   =', Vux, '_ kN') 
    print (' Vuy   =', Vuy, '_ kN')   
    print (' N° barra ref longt:', Nbx)
    print (' N° barra ref transv:', Nby)
    print ('')

# Refuerzo Longitudinal

    print ('Resultados refuerzo longitudinal')
    import f_pcal as pcalx
    pcalx = pcalx.f_pcal (fc, fy, Mux, b, d)
    print (" pcal  = %.4f" %pcalx)
    rtxt11.delete(0, 'end')
    rtxt11.insert (0, "{:.4f}".format(pcalx))

    Asx = pcalx*b*d
    print (" As    = %.2f" %Asx, '_ cm²')
    rtxt12.delete(0, 'end')
    rtxt12.insert (0, "{:.2f}".format (Asx))

    import f_Asmin as Asminx
    Asminx = Asminx.f_Asmin(pmin, b, h)
    print (" Asmin = %.2f" %Asminx, '_ cm²')
    rtxt13.delete(0, 'end')
    rtxt13.insert (0, "{:.2f}".format (Asminx))

    Asreqx = max(Asminx, Asx)
    print (" Asreq = %.2f" %Asreqx, '_ cm²')
    rtxt14.delete(0, 'end')
    rtxt14.insert (0, "{:.2f}".format (Asreqx))

    import f_Asb as Asbx
    Asbx = Asbx.f_Asb(Nbx)
    import math
    Cantx = math.ceil(Asreqx/Asbx)
    Ascolx = Asbx*Cantx

    print ('') 
    print('Distribución:')
    sx = b/(Asreqx/Asbx)
    if (sx > 30):
        sx = 30
    else:
        sx = sx
    print (' barras N°', "{:.0f}".format(Nbx), "C/ %.1f" %sx, ' cm')
    distx = str(Cantx) + ' barras N° ' +  str("{:.0f}".format(Nbx)) + " C/ " + "{:.1f}".format(sx) + ' cm' 
    rtxt15.delete(0, 'end')
    rtxt15.insert (0, distx)   

    print ('')
    print('Chequeo momento nominal:')
    print (" Ascol = %.2f" %Ascolx, '_ cm²')
    rtxt17.delete(0, 'end')
    rtxt17.insert (0, "{:.2f}".format (Ascolx))  
    import f_Mn as Mnx
    Mnx = Mnx.f_Mn(fc, fy, Ascolx, b, d)
    print (" øMn    = %.0f" %Mnx, '_ kN.cm')
    rtxt18.delete(0, 'end')
    rtxt18.insert (0, "{:.0f}".format (Mnx))

    if (Mux < Mnx):
        print (' Mu < øMn, Ok')   
        opc1 = 'Ok'
        rtxt19.delete(0, 'end')
        rtxt19.insert (0, opc1) 
    else:
        print ('Mu < øMn, No cumple')
        opc2 = ' No cumple'
        rtxt19.delete(0, 'end')
        rtxt19.insert (0, opc2)

# Refuerzo transversal
    print ('')
    print ('Resultados refuerzo transversal')
    import f_pcal as pcaly
    pcaly = pcaly.f_pcal (fc, fy, Muy, b, d)
    print (" pcal  = %.4f" %pcaly)
    rtxt51.delete(0, 'end')
    rtxt51.insert (0, "{:.4f}".format(pcaly))

    Asy = pcaly*b*d
    print (" As    = %.2f" %Asy, '_ cm²')
    rtxt52.delete(0, 'end')
    rtxt52.insert (0, "{:.2f}".format (Asy))

    import f_Asmin as Asminy
    Asminy = Asminy.f_Asmin(pmin, b, h)
    print (" Asmin = %.2f" %Asminy, '_ cm²')
    rtxt53.delete(0, 'end')
    rtxt53.insert (0, "{:.2f}".format (Asminy))

    Asreqy = max(Asminy, Asy)
    print (" Asreq = %.2f" %Asreqy, '_ cm²')
    rtxt54.delete(0, 'end')
    rtxt54.insert (0, "{:.2f}".format (Asreqy))

    import f_Asb as Asby
    Asby = Asby.f_Asb(Nby)
    import math
    Canty = math.ceil(Asreqy/Asby)
    Ascoly = Asby*Canty

    print ('') 
    print('Distribución:')
    sy = b/(Asreqy/Asby)
    if (sy > 30):
        sy = 30
    else:
        sy = sy
    print (' barras N°', "{:.0f}".format(Nby), "C/ %.1f" %sy, ' cm')
    disty = str(Canty) + ' barras N° ' +  str("{:.0f}".format(Nby)) + " C/ " + "{:.1f}".format(sy) + ' cm' 
    rtxt55.delete(0, 'end')
    rtxt55.insert (0, disty)   

    print ('')
    print('Chequeo momento nominal:')
    print (" Ascol = %.2f" %Ascoly, '_ cm²')
    rtxt57.delete(0, 'end')
    rtxt57.insert (0, "{:.2f}".format (Ascoly))  
    import f_Mn as Mny
    Mny = Mny.f_Mn(fc, fy, Ascoly, b, d)
    print (" øMn    = %.0f" %Mny, '_ kN.cm')
    rtxt58.delete(0, 'end')
    rtxt58.insert (0, "{:.0f}".format (Mny))

    if (Muy < Mny):
        print (' Mu < øMn, Ok')   
        opc1 = 'Ok'
        rtxt59.delete(0, 'end')
        rtxt59.insert (0, opc1) 
    else:
        print ('Mu < øMn, No cumple')
        opc2 = ' No cumple'
        rtxt59.delete(0, 'end')
        rtxt59.insert (0, opc2)

# Chequeo cortante simple dirección x

    print ('')
    print('Resistencia del concreto dirección x:')       
    import f_Vc as Vcx
    Vcx = Vcx.f_Vc(fiv, fc, b, d)
    print (" øVc    = %.2f" %Vcx, '_ kN')
    rtxt21.delete(0, 'end')
    rtxt21.insert (0, "{:.1f}".format (Vcx))

    if (Vux < Vcx):
        print (' Vu < øVc, Ok')   
        opc1 = 'Ok'
        rtxt22.delete(0, 'end')
        rtxt22.insert (0, opc1) 
    else:
        print (' Vu < øVc, No cumple')
        opc2 = 'No cumple'
        rtxt22.delete(0, 'end')
        rtxt22.insert (0, opc2)

# Chequeo cortante simple dirección y

    print ('')
    print('Resistencia del concreto dirección y:')       
    import f_Vc as Vcy
    Vcy = Vcy.f_Vc(fiv, fc, b, d)
    print (" øVc    = %.2f" %Vcy, '_ kN')
    rtxt61.delete(0, 'end')
    rtxt61.insert (0, "{:.1f}".format (Vcy))

    if (Vuy < Vcy):
        print (' Vu < øVc, Ok')   
        opc1 = 'Ok'
        rtxt62.delete(0, 'end')
        rtxt62.insert (0, opc1) 
    else:
        print (' Vu < øVc, No cumple')
        opc2 = 'No cumple'
        rtxt62.delete(0, 'end')
        rtxt62.insert (0, opc2)
    print ('')
    print(' ☺ El diseño a finalizado con exito ☺ ')

def reporte():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    book = load_workbook('D:\IEB_LOSAS\PlantillaLosa.xlsx')
    sheet = book.active

    fc = float (txt2.get())
    fy = float (txt3.get())

    b = float (txt4.get())
    sheet['C7'] = b
    sheet['C23'] = b*10
    sheet['D23'] = b*10

    h = float (txt5.get())
    sheet['C5'] = h

    r = float (txt6.get())
    pmin = float (txt7.get())
    fiv = float (txt8.get())

    Mux = float (txt9.get())
    sheet['C10'] = Mux

    Muy = float (txt10.get())
    sheet['D10'] = Muy
    
    Vux = float (txt11.get())
    sheet['C24'] = Vux
    
    Vuy = float (txt12.get())
    sheet['D24'] = Vuy

    Nbx = float (txt13.get())
    sheet['C15'] = Nbx

    Nby = float (txt14.get())
    sheet['D15'] = Nby

    d = h - r
    sheet['C6'] = d
    sheet['C22'] = d*10
    sheet['D22'] = d*10

    # Cálculo del acero

    import f_pcal as pcalx
    pcalx = pcalx.f_pcal (fc, fy, Mux, b, d)
    sheet['C11'] =  "%.4f" %pcalx

    import f_pcal as pcaly
    pcaly = pcaly.f_pcal (fc, fy, Muy, b, d)
    sheet['D11'] =  "%.4f" %pcaly

    Asx = pcalx*b*d
    sheet['C12'] =  "%.2f" %Asx

    Asy = pcaly*b*d
    sheet['D12'] =  "%.2f" %Asy

    import f_Asmin as Asminx
    Asminx = Asminx.f_Asmin(pmin, b, h)
    sheet['C13'] =  "%.2f" %Asminx

    import f_Asmin as Asminy
    Asminy = Asminy.f_Asmin(pmin, b, h)
    sheet['D13'] =  "%.2f" %Asminy

    Asreqx = max(Asminx, Asx)
    sheet['C14'] =  "%.2f" %Asreqx

    Asreqy = max(Asminy, Asy)
    sheet['D14'] =  "%.2f" %Asreqy

    import f_Asb as Asbx
    Asbx = Asbx.f_Asb(Nbx)
    import math
    Cantx = math.ceil(Asreqx/Asbx)
    sx = b/(Asreqx/Asbx)
    if (sx > 30):
        sx = 30
        sheet['C16'] ="%.2f" %sx
    else:
        sx = sx
        sheet['C16'] ="%.2f" %sx

    import f_Asb as Asby
    Asby = Asby.f_Asb(Nby)
    import math
    Canty = math.ceil(Asreqy/Asby)
    sy = b/(Asreqy/Asby)
    if (sy > 30):
        sy = 30
        sheet['C16'] ="%.2f" %sy
    else:
        sy = sy
        sheet['D16'] ="%.2f" %sy

    # Chequeo momento nominal 
    
    Ascolx = Asbx*Cantx
    Ascoly = Asby*Canty

    import f_Mn as Mnx
    Mnx = Mnx.f_Mn(fc, fy, Ascolx, b, d)
    sheet['C18'] =  "%.0f" %Mnx

    if (Mux < Mnx):
        opc1 = 'Ok'
        sheet['C19'] =  opc1        
    else:
        opc2 = 'No cumple'
        sheet['C19'] =  opc2

    import f_Mn as Mny
    Mny = Mny.f_Mn(fc, fy, Ascoly, b, d)
    sheet['D18'] =  "%.0f" %Mny

    if (Muy < Mny):
        opc1 = 'Ok'
        sheet['D19'] =  opc1        
    else:
        opc2 = 'No cumple'
        sheet['D19'] =  opc2

    # Chequeo cortante simple 
        
    import f_Vc as Vcx
    Vcx = Vcx.f_Vc(fiv, fc, b, d)
    sheet['C25'] =  "%.0f" %Vcx

    if (Vux < Vcx):
        opc1 = 'Ok'
        sheet['C26'] =  opc1 

    else:
        opc2 = 'No cumple'
        sheet['C26'] =  opc2 
        
    import f_Vc as Vcy
    Vcy = Vcy.f_Vc(fiv, fc, b, d)
    sheet['D25'] =  "%.0f" %Vcy

    if (Vuy < Vcy):
        opc1 = 'Ok'
        sheet['D26'] =  opc1 

    else:
        opc2 = 'No cumple'
        sheet['D26'] =  opc2 

    book.save ('D:\IEB_REPORTES\Diseño Losa.xlsx')

#Función para borrar celdas  

def Borrar ():

    txt2.delete(0, 'end')
    txt3.delete(0, 'end')
    txt4.delete(0, 'end')
    txt5.delete(0, 'end')
    txt6.delete(0, 'end')
    txt7.delete(0, 'end')
    txt8.delete(0, 'end')
    txt9.delete(0, 'end')
    txt10.delete(0, 'end')
    txt11.delete(0, 'end')
    txt12.delete(0, 'end')
    txt13.delete(0, 'end')
    txt14.delete(0, 'end')
    rtxt11.delete(0, 'end')
    rtxt12.delete(0, 'end')
    rtxt13.delete(0, 'end')
    rtxt14.delete(0, 'end')
    rtxt15.delete(0, 'end')
    rtxt17.delete(0, 'end')
    rtxt18.delete(0, 'end')
    rtxt19.delete(0, 'end')   
    rtxt21.delete(0, 'end')   
    rtxt22.delete(0, 'end')
    rtxt51.delete(0, 'end')
    rtxt52.delete(0, 'end')
    rtxt53.delete(0, 'end')
    rtxt54.delete(0, 'end')
    rtxt55.delete(0, 'end')
    rtxt57.delete(0, 'end')
    rtxt58.delete(0, 'end')
    rtxt59.delete(0, 'end')   
    rtxt61.delete(0, 'end')   
    rtxt62.delete(0, 'end')

# Cuadro 1 Datos iniciales

etiqueta1 = Label(ventana, text = "DATOS INICIALES", font='Helvetica 8 bold')
etiqueta1.pack()
etiqueta1.place(x=10, y=10, width=200, height=30)

etiqueta2 = Label(ventana, text = "fc. Resistencia del concreto _ MPa")
etiqueta2.pack()
etiqueta2.place(x=10, y=40, width=200, height=20)
txt2 = Entry(ventana, justify=CENTER)
txt2.place(x=215, y=40, width=100, height=20)

etiqueta3 = Label(ventana, text = "fy. Fluencia del acero _ MPa")
etiqueta3.pack()
etiqueta3.place(x=10, y=70, width=200, height=20)
txt3 = Entry(ventana, justify=CENTER)
txt3.place(x=215, y=70, width=100, height=20)

etiqueta4 = Label(ventana, text = "b. Base de la sección _ cm")
etiqueta4.pack()
etiqueta4.place(x=10, y=100, width=200, height=20)
txt4 = Entry(ventana, justify=CENTER)
txt4.place(x=215, y=100, width=100, height=20)

etiqueta5 = Label(ventana, text = "h. Altura de la sección _ cm")
etiqueta5.pack()
etiqueta5.place(x=10, y=130, width=200, height=20)
txt5 = Entry(ventana, justify=CENTER)
txt5.place(x=215, y=130, width=100, height=20)

etiqueta6 = Label(ventana, text = "r. Recubrimiento del refuerzo _ cm")
etiqueta6.pack()
etiqueta6.place(x=10, y=160, width=200, height=20)
txt6 = Entry(ventana, justify=CENTER)
txt6.place(x=215, y=160, width=100, height=20)

etiqueta7 = Label(ventana, text = "pmín. Cuantía mínima")
etiqueta7.pack()
etiqueta7.place(x=10, y=190, width=200, height=20)
txt7 = Entry(ventana, justify=CENTER)
txt7.place(x=215, y=190, width=100, height=20)

etiqueta8 = Label(ventana, text = "øv. Coeficiente de fricción")
etiqueta8.pack()
etiqueta8.place(x=10, y=220, width=200, height=20)
txt8 = Entry(ventana, justify=CENTER)
txt8.place(x=215, y=220, width=100, height=20)

etiqueta9 = Label(ventana, text = "Mux. Momento de diseño _ kN.cm")
etiqueta9.pack()
etiqueta9.place(x=10, y=250, width=200, height=20)
txt9 = Entry(ventana, justify=CENTER)
txt9.place(x=215, y=250, width=100, height=20)

etiqueta10 = Label(ventana, text = "Muy. Momento de diseño _ kN.cm")
etiqueta10.pack()
etiqueta10.place(x=10, y=280, width=200, height=20)
txt10 = Entry(ventana, justify=CENTER)
txt10.place(x=215, y=280, width=100, height=20)

etiqueta11 = Label(ventana, text = "Vux. Cortante de diseño _ kN")
etiqueta11.pack()
etiqueta11.place(x=10, y=310, width=200, height=20)
txt11 = Entry(ventana, justify=CENTER)
txt11.place(x=215, y=310, width=100, height=20)

etiqueta12 = Label(ventana, text = "Vuy. Cortante de diseño _ kN")
etiqueta12.pack()
etiqueta12.place(x=10, y=340, width=200, height=20)
txt12 = Entry(ventana, justify=CENTER)
txt12.place(x=215, y=340, width=100, height=20)

etiqueta13 = Label(ventana, text = "N°. barra ref longitudinal")
etiqueta13.pack()
etiqueta13.place(x=10, y=370, width=200, height=20)

# Nbarra=[2,3,4,5,6,7,8]
# var = StringVar(ventana)
# var.set(Nbarra[0])
# opc = OptionMenu(ventana, var, *Nbarra)
# opc.place(x=215, y=370, width=100, height=20)

txt13 = Entry(ventana, justify=CENTER)
txt13.place(x=215, y=370, width=100, height=20)

etiqueta14 = Label(ventana, text = "N°. barra ref transversal")
etiqueta14.pack()
etiqueta14.place(x=10, y=400, width=200, height=20)
txt14 = Entry(ventana, justify=CENTER)
txt14.place(x=215, y=400, width=100, height=20)

# Cuadro de resultados refuerzo

retiqueta10 = Label(ventana, text = "RESULTADOS", font='Helvetica 8 bold')
retiqueta10.pack()
retiqueta10.place(x=400, y=10, width=200, height=20)

retiqueta01 = Label(ventana, text = "Dir. Longitudinal", font='Helvetica 7 bold')
retiqueta01.pack()
retiqueta01.place(x=605, y=10, width=80, height=20)

retiqueta02 = Label(ventana, text = "Dir. Transversal", font='Helvetica 7 bold')
retiqueta02.pack()
retiqueta02.place(x=770, y=10, width=80, height=20)

retiqueta11 = Label(ventana, text = "preq. Cuatía requerida")  #Esta cuantia es igual a la cuantia calculada: pcal
retiqueta11.pack()
retiqueta11.place(x=400, y=40, width=200, height=30)
rtxt11 = Entry(ventana, justify=CENTER)
rtxt11.place(x=605, y=40, width=70, height=20)
rtxt11.config(bg = "#ecf0f1")
rtxt51 = Entry(ventana, justify=CENTER)
rtxt51.place(x=770, y=40, width=70, height=20)
rtxt51.config(bg = "#ecf0f1")

retiqueta12 = Label(ventana, text = "Asreq. Acero requerido _ cm²") # Este acero es el obtenido con al cuantía calculada: As caculado
retiqueta12.pack()
retiqueta12.place(x=400, y=70, width=200, height=20)
rtxt12 = Entry(ventana, justify=CENTER)
rtxt12.place(x=605, y=70, width=70, height=20)
rtxt12.config(bg = "#ecf0f1")
rtxt52 = Entry(ventana, justify=CENTER)
rtxt52.place(x=770, y=70, width=70, height=20)
rtxt52.config(bg = "#ecf0f1")

retiqueta13 = Label(ventana, text = "Asmín. Acero mínimo _ cm²")
retiqueta13.pack()
retiqueta13.place(x=400, y=100, width=200, height=20)
rtxt13 = Entry(ventana, justify=CENTER)
rtxt13.place(x=605, y=100, width=70, height=20)
rtxt13.config(bg = "#ecf0f1")
rtxt53 = Entry(ventana, justify=CENTER)
rtxt53.place(x=770, y=100, width=70, height=20)
rtxt53.config(bg = "#ecf0f1")

retiqueta14 = Label(ventana, text = "Ascol. Acero colocado _ cm²") # Este acero es igual al  Asreq. Acero requerido
retiqueta14.place(x=400, y=130, width=200, height=20)
rtxt14 = Entry(ventana, justify=CENTER)
rtxt14.place(x=605, y=130, width=70, height=20)
rtxt14.config(bg = "#ecf0f1")
rtxt54 = Entry(ventana, justify=CENTER)
rtxt54.place(x=770, y=130, width=70, height=20)
rtxt54.config(bg = "#ecf0f1")

retiqueta15 = Label(ventana, text = "Distribución del refurzo")
retiqueta15.pack()
retiqueta15.place(x=400, y=160, width=200, height=20)
rtxt15 = Entry(ventana, justify=CENTER)
rtxt15.place(x=605, y=160, width=140, height=20)
rtxt15.config(bg = "#ecf0f1")
rtxt55 = Entry(ventana, justify=CENTER)
rtxt55.place(x=770, y=160, width=140, height=20)
rtxt55.config(bg = "#ecf0f1")

# Chequeo

retiqueta16 = Label(ventana, text = "Chequeo momento nominal", font='Helvetica 8 bold')
retiqueta16.pack()
retiqueta16.place(x=400, y=190, width=200, height=20)

retiqueta03 = Label(ventana, text = "Mn Dir x", font='Helvetica 7 bold')
retiqueta03.pack()
retiqueta03.place(x=605, y=190, width=45, height=20)

retiqueta04 = Label(ventana, text = "Mn. Dir y", font='Helvetica 7 bold')
retiqueta04.pack()
retiqueta04.place(x=770, y=190, width=45, height=20)

retiqueta17 = Label(ventana, text = "Ascol: Acero colocado real _ cm²") # Este acero es igual al  Ascol Acero colocado , distribución final
retiqueta17.pack()
retiqueta17.place(x=400, y=230, width=200, height=20)
rtxt17 = Entry(ventana, justify=CENTER)
rtxt17.place(x=605, y=230, width=70, height=20)
rtxt17.config(bg = "#ecf0f1")
rtxt57 = Entry(ventana, justify=CENTER)
rtxt57.place(x=770, y=230, width=70, height=20)
rtxt57.config(bg = "#ecf0f1")

retiqueta18 = Label(ventana, text = "øMn. Momento nominal _ kN.cm")
retiqueta18.pack()
retiqueta18.place(x=400, y=260, width=200, height=20)
rtxt18 = Entry(ventana, justify=CENTER)
rtxt18.place(x=605, y=260, width=70, height=20)
rtxt18.config(bg = "#ecf0f1")
rtxt58 = Entry(ventana, justify=CENTER)
rtxt58.place(x=770, y=260, width=70, height=20)
rtxt58.config(bg = "#ecf0f1")

retiqueta19 = Label(ventana, text = "Mu < øMn")
retiqueta19.pack()
retiqueta19.place(x=400, y=290, width=200, height=20)
rtxt19 = Entry(ventana, justify=CENTER)
rtxt19.place(x=605, y=290, width=70, height=20)
rtxt19.config(bg = "#ecf0f1")
rtxt59 = Entry(ventana, justify=CENTER)
rtxt59.place(x=770, y=290, width=70, height=20)
rtxt59.config(bg = "#ecf0f1")

retiqueta20 = Label(ventana, text = "Chequeo resistencia del concreto", font='Helvetica 8 bold')
retiqueta20.pack()
retiqueta20.place(x=400, y=320, width=200, height=20)

retiqueta05 = Label(ventana, text = "Vu Dir x", font='Helvetica 7 bold')
retiqueta05.pack()
retiqueta05.place(x=605, y=320, width=45, height=20)

retiqueta06 = Label(ventana, text = "Vu. Dir y", font='Helvetica 7 bold')
retiqueta06.pack()
retiqueta06.place(x=770, y=320, width=45, height=20)

retiqueta21 = Label(ventana, text = "øVc. Resistencia del concreto _ kN")
retiqueta21.pack()
retiqueta21.place(x=400, y=350, width=200, height=20)
rtxt21 = Entry(ventana, justify=CENTER)
rtxt21.place(x=605, y=350, width=70, height=20)
rtxt21.config(bg = "#ecf0f1")
rtxt61 = Entry(ventana, justify=CENTER)
rtxt61.place(x=770, y=350, width=70, height=20)
rtxt61.config(bg = "#ecf0f1")

retiqueta22 = Label(ventana, text = "Vu < øVc")
retiqueta22.pack()
retiqueta22.place(x=400, y=380, width=200, height=20)
rtxt22 = Entry(ventana, justify=CENTER)
rtxt22.place(x=605, y=380, width=70, height=20)
rtxt22.config(bg = "#ecf0f1")
rtxt62 = Entry(ventana, justify=CENTER)
rtxt62.place(x=770, y=380, width=70, height=20)
rtxt62.config(bg = "#ecf0f1")

# Botones

boton1 = Button(ventana, text = 'Calcular', command=calcular)
boton1.pack()
boton1.place(x=185, y=450, width=150, height=20)

boton2 = Button(ventana, text = 'Guardar en xls', command=reporte)
boton2.pack()
boton2.place(x=385, y=450, width=150, height=20)

boton3 = Button(ventana, text = 'Borrar', command=Borrar)
boton3.pack()
boton3.place(x=585, y=450, width=150, height=20)

label = Label(ventana, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7')
label.pack()
label.place(x=347.5, y=485, width=225, height=10)

ventana.mainloop()